import csv
import itertools
import json
import os
from dataclasses import asdict
from pathlib import Path
from unittest.mock import patch

import pytest
import requests
import responses
from django.conf import settings
from django.contrib.gis.geos import Point
from openpyxl import Workbook
from rest_framework.exceptions import ValidationError

from iot import import_utils, models
from iot.import_utils import DuplicateReferenceError
from iot.serializers import Device2Serializer


def csv_to_workbook(path):
    """
    Convert a directory containing csv files into an excel workbook (one
    worksheet for each file).
    """
    workbook = Workbook()

    for file_entry in os.listdir(path):
        file_path = path / file_entry
        sheet = workbook.create_sheet(file_path.stem)
        with open(file_path, encoding='utf-8-sig') as f:
            for row in csv.reader(f, delimiter=';'):
                sheet.append(row)

    return workbook


def dict_to_workbook(value):
    """
    Convert a dict to a workbook. Each item in the dictionary is a sheet in the
    excel file. The key is the sheet name, the value is an iterable of iterables
    which are the rows and columns of the sheet.
    """
    workbook = Workbook()

    for sheet, rows in value.items():
        sheet = workbook.create_sheet(sheet)
        for row in rows:
            sheet.append(row)

    return workbook


class TestParse:

    @pytest.mark.parametrize("dir, parser, expected_location, expected_references,\
                             expected_project, expected_row_numbers", [
        (
            'iprox_single',
            import_utils.parse_iprox_xlsx,
            import_utils.Location(
                lat_long=None,
                postcode_house_number=import_utils.PostcodeHouseNumber('1011 PN', '3', 'III'),
                description='',
                regions=''
            ),
            ['7079-2296.0', '7079-2296.1'],
            [''],
            [1, 1],
        ),
        (
            'iprox_multiple',
            import_utils.parse_iprox_xlsx,
            import_utils.Location(
                lat_long=None,
                postcode_house_number=import_utils.PostcodeHouseNumber('1011 PN', '3', 'III'),
                description='',
                regions=''
            ),
            ['7079-2296.0', '7079-2297.0'],
            [''],
            [1, 2],
        ),
        (
            'bulk',
            import_utils.parse_bulk_xlsx,
            import_utils.Location(
                lat_long=import_utils.LatLong('52.3676', '4.9041'),
                postcode_house_number=None,
                description='',
                regions='AB;AC',
            ),
            ['7079-2296', '7079-2297'],
            ['Verkeershandhaving;Gemeente Amsterdam'],
            [1, 2],
        ),
    ])
    def test_parse_sensor_data(self, dir, parser, expected_location, expected_references,
                               expected_project, expected_row_numbers):
        """
        Each of the source files contains the same sensors, just in a different
        format, so we call the various parse functions and verify that
        """
        workbook = csv_to_workbook(Path(__file__).parent / 'data' / dir)
        actual = list(parser(workbook))

        expected_owner = import_utils.PersonData(
            organisation='Gemeente Amsterdam',
            email='p.ersoon@amsterdam.nl',
            telephone='06123456789',
            website='amsterdam.nl',
            first_name='Pieter',
            last_name_affix='',
            last_name='Ersoon',
        )

        # We expect the reference number and source to be different, so check the
        # fields individually
        expected = [
            import_utils.SensorData(
                reference=expected_references[0],
                owner=expected_owner,
                type='Optische / camera sensor',
                location=expected_location,
                datastream='Aantal schepen dat voorbij vaart',
                observation_goals=[import_utils.ObservationGoal(
                    observation_goal='Drukte en geluidsoverlast',
                    privacy_declaration='',
                    legal_ground=''
                )],
                themes='Veiligheid',
                contains_pi_data='Nee',
                active_until='06-07-2021',
                projects=expected_project,
                row_number=expected_row_numbers[0],
            ),
            import_utils.SensorData(
                reference=expected_references[1],
                owner=expected_owner,
                type='Chemiesensor',
                location=expected_location,
                datastream='Of er chemie is tussen mensen',
                observation_goals=[import_utils.ObservationGoal(
                    observation_goal='Vind ik gewoon leuk',
                    privacy_declaration='',
                    legal_ground='Bescherming vitale belangen betrokkene(n) of'
                                 ' van een andere natuurlijke persoon)'
                )],
                themes='Veiligheid;Lucht',
                contains_pi_data='Ja',
                active_until='01-01-2050',
                projects=[''],
                row_number=expected_row_numbers[1],
            ),
        ]

        assert list(map(asdict, actual)) == list(map(asdict, expected))

    @pytest.mark.parametrize(
        "fields",
        [
            [],
            ['non', 'sense'],
            import_utils.IPROX_FIELDS[:-1],
            import_utils.IPROX_FIELDS[:-1] + ['nonsense'],
        ]
    )
    def test_parse_prox_invalid_fields_should_be_detected(self, fields):
        # check that an incorrect number of fields can be correctly detected
        workbook = dict_to_workbook({'Sensorregistratie': [fields]})
        with pytest.raises(import_utils.InvalidIproxFields):
            list(import_utils.parse_iprox_xlsx(workbook))

    @pytest.mark.parametrize(
        "fields",
        [
            [],
            ['non', 'sense'],
            import_utils.BULK_PERSON_FIELDS[:-1],
            import_utils.BULK_PERSON_FIELDS[:-1] + ['nonsense'],
        ]
    )
    def test_parse_bulk_invalid_person_fields_should_be_detected(self, fields):
        # check that an incorrect number of fields can be correctly detected
        workbook = dict_to_workbook({'Uw gegevens': [[f, f] for f in fields]})
        with pytest.raises(import_utils.InvalidPersonFields):
            list(import_utils.parse_bulk_xlsx(workbook))

    @pytest.mark.parametrize(
        "fields",
        [
            [],
            ['non', 'sense'],
            import_utils.BULK_SENSOR_FIELDS[:-1],
            import_utils.BULK_SENSOR_FIELDS[:-1] + ['nonsense'],
        ]
    )
    def test_parse_bulk_invalid_sensor_fields_should_be_detected(self, fields):
        # check that an incorrect number of fields can be correctly detected
        person_fields = [[f, f] for f in import_utils.BULK_PERSON_FIELDS]
        workbook = dict_to_workbook({'Uw gegevens': person_fields, 'Sensorregistratie': [fields]})
        with pytest.raises(import_utils.InvalidSensorFields):
            list(import_utils.parse_bulk_xlsx(workbook))

    @pytest.mark.parametrize("dir, parser", [
        (
            'empty_rows_columns_iprox',
            import_utils.parse_iprox_xlsx,
        ),
        (
            'empty_rows_columns_bulk',
            import_utils.parse_bulk_xlsx,
        ),
    ])
    def test_empty_rows_and_columns_should_be_skipped(self, dir, parser):
        # verify that empty rows and columns after the expected data is ignored
        workbook = csv_to_workbook(Path(__file__).parent / 'data' / dir)
        assert len(list(parser(workbook))) == 1

    def test_iprox_five_sensors(self):
        # verify that is possible to import the max number of sensors
        workbook = csv_to_workbook(Path(__file__).parent / 'data' / 'iprox_five_sensors')
        assert len(list(import_utils.parse_iprox_xlsx(workbook))) == 5


@pytest.fixture
def person_data():
    return import_utils.PersonData(
        organisation='Gemeente Amsterdam',
        email='p.er.soon@amsterdam.nl',
        telephone='06123456789',
        website='amsterdam.nl',
        first_name='Piet',
        last_name_affix='Er',
        last_name='Soon',
    )


@pytest.mark.django_db
class TestImportPerson:

    @property
    def actual(self):
        fields = 'organisation', 'email', 'telephone', 'website', 'name'
        return list(models.Person2.objects.values(*fields).order_by('id').all())

    expected = {
        'organisation': 'Gemeente Amsterdam',
        'email': 'p.er.soon@amsterdam.nl',
        'telephone': '06123456789',
        'website': 'amsterdam.nl',
        'name': 'Piet Er Soon',
    }

    def test_import_person(self, person_data):
        import_utils.import_person(person_data)
        assert self.actual == [self.expected]

    def test_import_person_should_be_idempotent(self, person_data):
        # check that we only import the same person once
        import_utils.import_person(person_data)
        import_utils.import_person(person_data)
        assert self.actual == [self.expected]

    def test_import_person_should_be_case_insensitive(self, person_data):
        # check that matching by email address is case insensitive
        import_utils.import_person(person_data)
        person_data.email = 'p.Er.SoOn@AmStErDaM.nL'
        import_utils.import_person(person_data)
        assert self.actual == [self.expected]

    def test_import_person_should_update_values(self, person_data):
        # check that a second import of the same person updates values
        import_utils.import_person(person_data)
        person_data.telephone = '0201234567890'
        import_utils.import_person(person_data)
        assert self.actual == [dict(self.expected, telephone='0201234567890')]

    def test_import_person_should_import_multiple_people(self, person_data):
        # check that difference email addresses result in different people
        import_utils.import_person(person_data)
        person_data.email = 'p.er.soon@rotterdam.nl'
        import_utils.import_person(person_data)
        assert self.actual == [self.expected, dict(self.expected, email='p.er.soon@rotterdam.nl')]


@pytest.fixture
def sensor_data(person_data):
    return import_utils.SensorData(
        owner=person_data,
        reference='1234',
        type='Chemiesensor',
        location=import_utils.Location(
            lat_long=None,
            postcode_house_number=None,
            description='Somewhere over the rainbow',
            regions='',
        ),
        datastream='water',
        observation_goals=[import_utils.ObservationGoal(
            observation_goal='Nare bedoelingen',
            privacy_declaration='https://amsterdam.nl/privacy',
            legal_ground='Publieke taak'
        )],
        themes=settings.IPROX_SEPARATOR.join([
            'Mobiliteit: auto',
            'Mobiliteit: fiets',
        ]),
        contains_pi_data='Ja',
        active_until='05-05-2050',
        projects=[
            'Gemeente Amsterdam;Verkeers',
            'Noord Amsterdam;Zuid Amsterdam'
        ],
    )


@pytest.mark.django_db
class TestImportSensor:

    @property
    def actual(self):
        return [
            Device2Serializer(device).data
            for device in models.Device2.objects.all().order_by('id')
        ]

    expected = {
        'active_until': '2050-05-05',
        'contains_pi_data': True,
        'datastream': 'water',
        'location': None,
        'location_description': "Somewhere over the rainbow",
        'observation_goals': [
            {
                'observation_goal': 'Nare bedoelingen',
                'privacy_declaration': 'https://amsterdam.nl/privacy',
                'legal_ground': 'Publieke taak'
            }
        ],
        'owner': {
            'name': 'Piet Er Soon',
            'email': 'p.er.soon@amsterdam.nl',
            'organisation': 'Gemeente Amsterdam',
        },
        'regions': [],
        'themes': [
            'Mobiliteit: auto',
            'Mobiliteit: fiets',
        ],
        'type': 'Chemiesensor',
        'reference': '1234',
        'project_paths': [
            ['Gemeente Amsterdam', 'Verkeers'],
            ['Noord Amsterdam', 'Zuid Amsterdam']
        ],
    }

    def test_import_sensor(self, sensor_data):
        # Basic check for importing sensor data
        owner = import_utils.import_person(sensor_data.owner)
        import_utils.import_sensor(sensor_data, owner)
        assert self.actual == [self.expected]

    def test_import_sensor_should_be_idempotent(self, sensor_data):
        # Check that importing the same data twice is idempotent
        owner = import_utils.import_person(sensor_data.owner)
        import_utils.import_sensor(sensor_data, owner)
        import_utils.import_sensor(sensor_data, owner)
        assert self.actual == [self.expected]

    def test_import_sensor_should_update_values(self, sensor_data):
        # check that a second import of the same sensor updates values
        owner = import_utils.import_person(sensor_data.owner)
        import_utils.import_sensor(sensor_data, owner)
        sensor_data.observation_goals[0].privacy_declaration = 'http://rotterdam.nl/privacy'
        import_utils.import_sensor(sensor_data, owner)
        assert self.actual == [
            dict(self.expected, observation_goals=[
                {
                    'observation_goal': 'Nare bedoelingen',
                    'privacy_declaration': 'https://amsterdam.nl/privacy',
                    'legal_ground': 'Publieke taak'
                },
                {
                    'observation_goal': 'Nare bedoelingen',
                    'privacy_declaration': 'http://rotterdam.nl/privacy',
                    'legal_ground': 'Publieke taak'
                }
            ]
            )
        ]

    def test_import_sensor_should_import_multiple_sensors(self, sensor_data):
        # check that we can import multiple sensors
        owner = import_utils.import_person(sensor_data.owner)
        import_utils.import_sensor(sensor_data, owner)
        sensor_data.reference = '2468'
        import_utils.import_sensor(sensor_data, owner)
        assert self.actual == [self.expected, dict(self.expected, reference='2468')]

    def test_import_sensor_location(self, sensor_data):
        # check that the location is imported correctly
        owner = import_utils.import_person(sensor_data.owner)
        sensor_data.location.lat_long = import_utils.LatLong(latitude=52.3676, longitude=4.9041)
        import_utils.import_sensor(sensor_data, owner)
        location = {"latitude": 52.3676, "longitude": 4.9041}
        assert self.actual == [
            dict(
                self.expected,
                location_description='Somewhere over the rainbow',
                location=location
            )
        ]

    def test_import_sensor_other_type(self, sensor_data):
        # check that we can import a "other" sensor type
        owner = import_utils.import_person(sensor_data.owner)
        sensor_data.type = "Midi-chlorian teller"
        import_utils.import_sensor(sensor_data, owner)
        assert self.actual == [dict(self.expected, type="Overig")]

    def test_import_sensor_other_themes(self, sensor_data):
        # check that we can import some "other" sensor themes
        owner = import_utils.import_person(sensor_data.owner)
        sensor_data.themes = "abc;123"
        import_utils.import_sensor(sensor_data, owner)
        assert self.actual == [dict(self.expected, themes=["Overig"])]

    def test_import_sensor_regions(self, sensor_data):
        # check that we can import as "mobile" sensor + the location_description is there
        owner = import_utils.import_person(sensor_data.owner)
        regions = ["Centrum", "Oost"]
        sensor_data.location.regions = "Centrum;Oost"
        import_utils.import_sensor(sensor_data, owner)
        expected = dict(
            self.expected,
            location_description='Somewhere over the rainbow',
            regions=regions
        )
        assert self.actual == [expected]

    def test_import_sensor_other_region(self, sensor_data):
        # check that we can import a "other" region and location_description is there.
        owner = import_utils.import_person(sensor_data.owner)
        sensor_data.location.regions = "Diemen"
        import_utils.import_sensor(sensor_data, owner)
        assert self.actual == [
            dict(
                self.expected,
                location_description='Somewhere over the rainbow',
                regions=["Diemen"]
            )
        ]

    def test_import_postcode_house_number(self, sensor_data):
        # check that we can import a location based on postcode and house number
        owner = import_utils.import_person(sensor_data.owner)
        sensor_data.location.postcode_house_number = import_utils.PostcodeHouseNumber("1111AA", 1)
        with patch('iot.import_utils.get_center_coordinates', lambda *_: Point(4.9041, 52.3676)):
            import_utils.import_sensor(sensor_data, owner)
        location = {"latitude": 52.3676, "longitude": 4.9041}
        assert self.actual == [
            dict(
                self.expected,
                location_description='Somewhere over the rainbow',
                location=location
            )
        ]

    def test_import_observation_goal(self, sensor_data):
        # check that we can import an observation_goal
        owner = import_utils.import_person(sensor_data.owner)
        sensor_data.observation_goals = [
            import_utils.ObservationGoal(
                observation_goal='my_observation_goal',
                privacy_declaration='my_privacy_declaration',
                legal_ground='my_legal_ground'
            )
        ]
        import_utils.import_sensor(sensor_data, owner)
        assert self.actual == self.actual == [
            dict(self.expected, observation_goals=[
                {
                    'observation_goal': 'my_observation_goal',
                    'privacy_declaration': 'my_privacy_declaration',
                    'legal_ground': 'my_legal_ground'
                }
            ]
            )
        ]

    @pytest.mark.parametrize("source", ["iprox", "bulk"])
    def test_duplicate_references_should_be_rejected(self, source):
        path = Path(__file__).parent / 'data' / f'{source}_duplicate_references'
        workbook = csv_to_workbook(path)
        errors = import_utils.import_xlsx(workbook)[0]
        expected = [
            DuplicateReferenceError('7079-2296.0', 2),
            DuplicateReferenceError('7079-2296.1', 2),
        ]
        assert errors == expected


@pytest.mark.django_db
class TestValidate:
    """
    Check that the validation function can report problematic values.
    """

    @pytest.mark.parametrize("value", [None, '', ' ', '  '])
    def test_invalid_sensor_type(self, sensor_data, value):
        sensor_data.type = value
        with pytest.raises(import_utils.InvalidSensorType):
            import_utils.validate_sensor(sensor_data)

    @pytest.mark.parametrize("value", [None, '', ' ', '  '])
    def test_invalid_themes(self, sensor_data, value):
        sensor_data.themes = value
        with pytest.raises(import_utils.InvalidThemes):
            import_utils.validate_sensor(sensor_data)

    @pytest.mark.parametrize("value", [None, '', 'onzin'])
    def test_invalid_latitude(self, sensor_data, value):
        sensor_data.location.lat_long = import_utils.LatLong(value, 4)
        with pytest.raises(import_utils.InvalidLatitude):
            import_utils.validate_sensor(sensor_data)

    @pytest.mark.parametrize("value", [None, '', 'onzin'])
    def test_invalid_longitude(self, sensor_data, value):
        sensor_data.location.lat_long = import_utils.LatLong(52, value)
        with pytest.raises(import_utils.InvalidLongitude):
            import_utils.validate_sensor(sensor_data)

    @pytest.mark.parametrize("value", [None, '', 'onzin', '1111', 'XX'])
    def test_invalid_postcode(self, sensor_data, value):
        sensor_data.location.postcode_house_number = import_utils.PostcodeHouseNumber(value, 1)
        with pytest.raises(import_utils.InvalidPostcode):
            import_utils.validate_sensor(sensor_data)

    @pytest.mark.parametrize("value", [None, '', 'Yes', True, False, 'true', 'false'])
    def test_invalid_contains_pi_data(self, sensor_data, value):
        sensor_data.contains_pi_data = value
        with pytest.raises(import_utils.InvalidContainsPiData):
            import_utils.validate_sensor(sensor_data)

    @pytest.mark.parametrize("value", [None, ''])
    def test_invalid_legal_ground(self, sensor_data, value):
        sensor_data.observation_goals[0].legal_ground = value
        with pytest.raises(import_utils.InvalidLegalGround):
            import_utils.validate_sensor(sensor_data)

    @pytest.mark.parametrize("value", [None, '', '5 Nov 1605', '1999/12/31', 'onzin'])
    def test_invalid_date(self, sensor_data, value):
        sensor_data.active_until = value
        with pytest.raises(import_utils.InvalidDate):
            import_utils.validate_sensor(sensor_data)

    @pytest.mark.parametrize("value", [None, ''])
    def test_invalid_privacy_declaration(self, sensor_data, value):
        # don't accept empty values when the sensor collects personal data
        sensor_data.contains_pi_data = 'Ja'
        sensor_data.observation_goals[0].privacy_declaration = value
        with pytest.raises(import_utils.InvalidPrivacyDeclaration):
            import_utils.validate_sensor(sensor_data)

    @pytest.mark.parametrize(
        "value,contains_pi_data",
        itertools.product(['nee', '-', 'n/a', 'n.v.t.'], ['Ja', 'Nee']),
    )
    def test_invalid_privacy_declaration_invalid_urls(self, sensor_data, value, contains_pi_data):
        # never accept invalid urls
        sensor_data.contains_pi_data = contains_pi_data
        sensor_data.observation_goals[0].privacy_declaration = value
        with pytest.raises(import_utils.InvalidPrivacyDeclaration):
            import_utils.validate_sensor(sensor_data)

    @pytest.mark.parametrize("value", [None, ''])
    def test_invalid_privacy_declaration_can_be_empty_if_no_pi_data(self, sensor_data, value):
        sensor_data.contains_pi_data = 'Nee'
        sensor_data.observation_goals[0].privacy_declaration = value
        try:
            import_utils.validate_sensor(sensor_data)
        except import_utils.InvalidPrivacyDeclaration:
            pytest.fail('Should not raise on empty declaration when no personal data collected')

    @pytest.mark.parametrize("value", [None, '', ' ', 'not-an-email'])
    def test_invalid_email(self, person_data, value):
        person_data.email = value
        with pytest.raises(ValidationError):
            person_data.validate()

    @pytest.mark.parametrize("value", [None, '', ' '])
    def test_invalid_first_name(self, person_data, value):
        person_data.first_name = value
        with pytest.raises(ValidationError):
            person_data.validate()

    @pytest.mark.parametrize("value", [None, '', ' '])
    def test_invalid_last_name(self, person_data, value):
        person_data.last_name = value
        with pytest.raises(ValidationError):
            person_data.validate()


DATA_DIR = Path(__file__).parent / 'data' / 'json'


@pytest.fixture()
def requests_mock():
    with responses.RequestsMock() as requests_mock:
        yield requests_mock


def mock_get(requests_mock, url, response):
    response = json.loads(response) if isinstance(response, str) else response
    requests_mock.add(responses.GET, url, json=response)
    requests.get(url)


@pytest.fixture(autouse=True)
def override_url_settings(settings):
    # make sure we don't accidentally call the actual API during tests
    # we should use the mocked values
    settings.ATLAS_POSTCODE_SEARCH = 'http://postcode'
    settings.ATLAS_ADDRESS_SEARCH = 'http://adres'


class TestGetCenterCoordinates:

    @pytest.fixture(autouse=True)
    def inject_requests_mock(self, requests_mock):
        self.requests_mock = requests_mock
        for path in os.listdir(DATA_DIR):
            for file in os.listdir(DATA_DIR / path):
                response = json.loads((DATA_DIR / path / file).read_text())
                query = file.replace(".json", "")
                url = f'http://{path}/?q={query}'
                mock_get(self.requests_mock, url, response)

    def test_house_number_expected_point(self):
        """
        Check that when the correct response is given we get the expected
        result for the postcode + house number.
        """
        actual = import_utils.get_center_coordinates('1015 BA', 1)
        assert tuple(actual) == (4.892287512614596, 52.37905673120624)

    def test_house_number_not_found_should_raise_exception(self):
        """
        Searching by herengracht%201&page=2.json is fuzzy, so for example
        Herengracht 11 will return results, even though there is no Herengracht
        11, in this case the closest text result is Herengracht 111. In this
        situation the best thing is to just produce an error and let the person
        who is doing the import figure out the best course of action.
        """
        with pytest.raises(import_utils.PostcodeSearchException):
            import_utils.get_center_coordinates('1015 BA', 11)

    def test_pagination_links_should_be_followed(self):
        """
        It may be that the herengracht we are looking for is not on the first page,
        for example when searching for "Herengracht 1" and we mean the
        Heregracht in Weesp.
        """
        actual = import_utils.get_center_coordinates('1382 AE', 1)
        assert tuple(actual) == (5.042682700063272, 52.30925075920674)

    @pytest.mark.parametrize("postcode", ['1016ht', '1016 ht', '1016 HT', '1016HT'])
    def test_postcode_should_be_normalized(self, postcode):
        """
        Check that the postcode is correctly normalized, if a postcode
        is given in lowercase it will not match the API.
        """
        actual = import_utils.get_center_coordinates(postcode, 579)
        assert tuple(actual) == (4.882909214312223, 52.36816000657591)


class TestGetCenterCoordinatesInvalidResponses:

    HERENGRACHT_POSTCODE_RESPONSE = json.dumps({"results": [{"straat": 'Herengracht'}]})

    @pytest.fixture(autouse=True)
    def inject_requests_mock(self, requests_mock):
        self.requests_mock = requests_mock

    @pytest.mark.parametrize('postcode_response, address_response', [
        ('{}', '{}'),
        ('{"results": null}', '{}'),
        ('{"results": []}', '{}'),
        ('{"results": [{}]}', '{}'),
        ('{"results": [{"straat": null}]}', '{}'),
        (HERENGRACHT_POSTCODE_RESPONSE, '{}'),
        (HERENGRACHT_POSTCODE_RESPONSE, '{"results": null}'),
        (HERENGRACHT_POSTCODE_RESPONSE, '{"results": []}'),
        (HERENGRACHT_POSTCODE_RESPONSE, '{"results": [{}]}'),
    ])
    def test_house_number_exception_should_be_raised_on_invalid_response(
        self,
        postcode_response: str,
        address_response: str,
    ):
        """
        When getting the center coordinates of a postcode with a house number,
        check that an exception is raised if the response (postcode or herengracht%201&page=2.json)
        from atlas is not in the expected form.
        """
        url = import_utils.get_postcode_url('1015 BA')
        mock_get(self.requests_mock, url, postcode_response)
        url = import_utils.get_address_url('Herengracht', 1)
        mock_get(self.requests_mock, url, address_response)

        with pytest.raises(import_utils.PostcodeSearchException):
            import_utils.get_center_coordinates('1015 BA', 1)
